def generate_bmw_dsr(input_file, output_file):
    import pandas as pd
    import re
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    INPUT_FILE = input_file
    OUTPUT_FILE = output_file
    VENDOR_NAME = "BMW AG"

    OUTPUT_COLUMNS = [
        "SHIPMENT NO.",
        "JOB NO.",
        "VENDOR / SUPPLIER NAME",
        "VENDOR/ SUPPLIER ADDRESS (COUNTRY OF ORIGIN)",
        "VESSEL NAME",
        "PORT OF LANDING",
        "B/L NO / MAWB NO",
        "B/L  DATE / MAWB DATE",
        "NO OF CARS",
        "ETA / SHIPMENT DATE",
        "VIN NO",
        "MODEL NO",
        "AG INVOICE NUMBER",
        "AG INVOICE DATE",
        "AG INVOICE VALUE",
        "COMMERCIAL INVOICE CURRENCY",
        "PORT OF DISCHARGE / PORT CODE",
        "BoE #",
        "BoE Date",
        "AD CODE",
        "ASSESABLE VALUE",
        "DUTY BANK PAYMENT",
        "DUTY LICENCE DEBIT",
        "TOTAL DUTY",
        "IGM NO & ITEM NO",
        "IGM DATE",
        "INWARD DATE",
        "OOC DATE",
        "RECEIPT OF INVOICE",
        "RECEIPT OF TYPE APPROVAL CERTIFICATE",
        "RECEIPT OF CERTIFICATE OF ORIGIN",
        "RECEIPT OF B/L",
        "SVB",
        "SHIFT",
        "DATE OF HANDOVER TO TRANSPORTER",
        "REMARK"
    ]

    # ---------------- READ INPUT ----------------
    df = pd.read_excel(INPUT_FILE, dtype=str)
    df.columns = df.columns.str.strip()
    df = df.fillna("")

    # ---------------- VIN EXTRACTION ----------------
    def extract_vin(text):
        match = re.search(r"VIN NO\s+([A-Z0-9]{17})", text)
        return match.group(1) if match else ""

    df["VIN_NO"] = df["Product Desc"].apply(extract_vin)

    def to_float(val):
                try:
                    return float(str(val).replace(",", ""))
                except:
                    return 0.0

    def split_unit_price(value):
        parts = str(value).strip().split()
        if len(parts) == 2:
            return parts[0], parts[1]
        return "", ""

    # ---------------- BUILD OUTPUT ----------------
    output_rows = []
    shipment_no = 0

    for job_no, group in df.groupby("Job No", sort=False):
        shipment_no += 1
        total_cars = len(group)
        first_row = True

        # VIN rows
        for _, row in group.iterrows():
            inv_value, inv_curr = split_unit_price(row["Unit Price"])
                
            duty_bank = to_float(row["Total Duty (INR)"])
            duty_licence = to_float(row["BCD Foregone"])
            calculated_total_duty = duty_bank + duty_licence

            output_rows.append({
                "SHIPMENT NO.": shipment_no if first_row else "",
                "JOB NO.": job_no,
                "VENDOR / SUPPLIER NAME": VENDOR_NAME,
                "VENDOR/ SUPPLIER ADDRESS (COUNTRY OF ORIGIN)": row["Country of Origin"],
                "VESSEL NAME": "",
                "PORT OF LANDING": "",
                "B/L NO / MAWB NO": row["MAWB/MBL No"],
                "B/L  DATE / MAWB DATE": "",
                "NO OF CARS": 1,
                "ETA / SHIPMENT DATE": "",
                "VIN NO": row["VIN_NO"],
                "MODEL NO": row["Product Desc"],
                "AG INVOICE NUMBER": row["Invoice No"],
                "AG INVOICE DATE": row["Invoice Date"],
                "AG INVOICE VALUE": inv_value,
                "COMMERCIAL INVOICE CURRENCY": inv_curr,
                "PORT OF DISCHARGE / PORT CODE": "",
                "BoE #": row["BE No"],
                "BoE Date": row["BE Date"],
                "AD CODE": "",
                "DUTY BANK PAYMENT": duty_bank,
                "DUTY LICENCE DEBIT": duty_licence,
                "ASSESABLE VALUE": row["Assessable Value (INR)"],
                "TOTAL DUTY": calculated_total_duty,
                "IGM NO & ITEM NO": "",
                "IGM DATE": "",
                "INWARD DATE": "",
                "OOC DATE": "",
                "RECEIPT OF INVOICE": "",
                "RECEIPT OF TYPE APPROVAL CERTIFICATE": "",
                "RECEIPT OF CERTIFICATE OF ORIGIN": "",
                "RECEIPT OF B/L": "",
                "SVB": "",
                "SHIFT": "",
                "DATE OF HANDOVER TO TRANSPORTER": "",
                "REMARK": ""
            })
            first_row = False

        # TOTAL row (single)
        output_rows.append({
            "SHIPMENT NO.": "",
            "JOB NO.": "",
            "VENDOR / SUPPLIER NAME": "",
            "VENDOR/ SUPPLIER ADDRESS (COUNTRY OF ORIGIN)": "",
            "VESSEL NAME": "",
            "PORT OF LANDING": "",
            "B/L NO / MAWB NO": "",
            "B/L  DATE / MAWB DATE": "",
            "NO OF CARS": total_cars,
            "ETA / SHIPMENT DATE": "",
            "VIN NO": "",
            "MODEL NO": "",
            "AG INVOICE NUMBER": "",
            "AG INVOICE DATE": "",
            "AG INVOICE VALUE": "",
            "COMMERCIAL INVOICE CURRENCY": "",
            "PORT OF DISCHARGE / PORT CODE": "",
            "BoE #": "",
            "BoE Date": "",
            "AD CODE": "",
            "DUTY BANK PAYMENT": "",
            "DUTY LICENCE DEBIT": "",
            "ASSESABLE VALUE": "",
            "TOTAL DUTY": "",
            "IGM NO & ITEM NO": "",
            "IGM DATE": "",
            "INWARD DATE": "",
            "OOC DATE": "",
            "RECEIPT OF INVOICE": "",
            "RECEIPT OF TYPE APPROVAL CERTIFICATE": "",
            "RECEIPT OF CERTIFICATE OF ORIGIN": "",
            "RECEIPT OF B/L": "",
            "SVB": "",
            "SHIFT": "",
            "DATE OF HANDOVER TO TRANSPORTER": "",
            "REMARK": ""
        })

    # ---------------- WRITE EXCEL ----------------
    out_df = pd.DataFrame(output_rows, columns=OUTPUT_COLUMNS)
    out_df.to_excel(OUTPUT_FILE, index=False)

    # ---------------- FORMATTING ----------------
    wb = load_workbook(OUTPUT_FILE)
    ws = wb.active

    headers = {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}

    # ---------------- HEADER FORMATTING ----------------

    header_alignment = Alignment(
        wrap_text=True,
        vertical="center",
        horizontal="center"
    )

    # Increase header row height slightly
    ws.row_dimensions[1].height = 35

    # Apply wrap + bold + center to header cells
    for cell in ws[1]:
        cell.alignment = header_alignment
        cell.font = Font(bold=True)

    # ---------------- AUTO COLUMN WIDTH ----------------

    for col_cells in ws.columns:
        max_length = 0
        col_letter = col_cells[0].column_letter

        for cell in col_cells:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        # Add padding so text is not cramped
        adjusted_width = min(max_length + 3, 45)
        ws.column_dimensions[col_letter].width = adjusted_width

    # ---------------- DATE COLUMN FORMATTING ----------------

    DATE_COLUMNS = [
        "AG INVOICE DATE",
        "BoE Date",
        "IGM DATE",
        "INWARD DATE",
        "OOC DATE",
        "B/L  DATE / MAWB DATE",
        "ETA / SHIPMENT DATE"
    ]

    date_format = "DD-MMM-YYYY"

    for col_name in DATE_COLUMNS:
        if col_name in headers:
            col_idx = headers[col_name]
            for row in ws.iter_rows(min_row=2):
                cell = row[col_idx - 1]
                if cell.value not in ("", None):
                    cell.number_format = date_format

    SHIPMENT_ROW_HEIGHT = 30   # adjust if needed (default Excel ~15)
    SHIPMENT_FONT_SIZE = 22

    shipment_col_idx = headers["SHIPMENT NO."]

    for row in ws.iter_rows(min_row=2):
        shipment_cell = row[shipment_col_idx - 1]
        if shipment_cell.value not in ("", None):
            ws.row_dimensions[shipment_cell.row].height = SHIPMENT_ROW_HEIGHT
            # Make shipment number bold & slightly larger
            shipment_cell.font = Font(
                bold=True,
                size=SHIPMENT_FONT_SIZE
            )

            # center vertically for visual balance
            shipment_cell.alignment = Alignment(
                vertical="center",
                horizontal="center"
            )

    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    bold_font = Font(bold=True)

    for row in ws.iter_rows(min_row=2):
        shipment_val = row[headers["SHIPMENT NO."] - 1].value
        cars_val = row[headers["NO OF CARS"] - 1].value

        if (shipment_val in ("", None)) and isinstance(cars_val, (int, float)) and cars_val > 1:
            row[headers["NO OF CARS"] - 1].fill = yellow_fill
            row[headers["NO OF CARS"] - 1].font = bold_font

    wb.save(OUTPUT_FILE)
